import datetime
import os
import pickle
import threading
from threading import Thread

import cv2
import face_recognition
import numpy as np
import openpyxl
from flask import (
    Flask,
    Response,
    flash,
    jsonify,
    redirect,
    render_template,
    request,
    send_file,
    url_for,
)
from tensorflow.keras.preprocessing.image import ImageDataGenerator, img_to_array
from tqdm import tqdm

app = Flask(__name__)
app.secret_key = 'your_secret_key_here'  # Required for flash messages


attendance_thread = None

def recognize_face(features, student_features):
    min_distance = float('inf')
    recognized_student = None

    for student, known_features in student_features.items():
        distance = face_recognition.face_distance(np.array([features]), np.array([known_features]))[0]
        if distance < min_distance:
            min_distance = distance
            recognized_student = student

    # Set a threshold for recognition (adjust as needed)
    threshold = 0.5
    if min_distance < threshold:
        return recognized_student
    else:
        return None

def load_registration_data(filename):
    wb = openpyxl.load_workbook(filename=filename, data_only=True)
    sheet = wb.active  # Assuming registration data is in the first sheet

    registration_data = {}
    for row in sheet.iter_rows(min_row=2):  # Skip the header row (row 1)
        name = row[1].value  # Assuming name is in the first column (column A)
        reg_no = row[2].value  # Assuming registration number is in the second column (column B)
        semester = row[6].value  # Column D for semester
        section = row[7].value  # Column E for section
        registration_data[name] = {'reg_no': reg_no, 'semester': semester, 'section': section}

    return registration_data

def write_to_excel(name, reg_no, timestamp, sheet, last_attendance_times, semester, section):
    # Convert timestamp to datetime object
    current_time = datetime.datetime.strptime(timestamp, "%d-%m-%Y %H:%M:%S")
    date = current_time.strftime("%d-%m-%Y")
    time =  current_time.strftime("%H:%M:%S")

    # Check if the person has already been recorded within the last 2 minutes
    if name in last_attendance_times and last_attendance_times[name] > current_time - datetime.timedelta(minutes=2):
        return  # Entry already exists within the last 2 minutes, skip writing

    # Update last attendance time for the person
    last_attendance_times[name] = current_time

    # Find the next empty row
    next_row = 1
    while sheet.cell(row=next_row, column=1).value is not None:
        next_row += 1

    # Write the data to the next empty row 
    sheet.cell(row=next_row, column=1).value = next_row  # Serial number
    sheet.cell(row=next_row, column=2).value = name
    sheet.cell(row=next_row, column=3).value = reg_no
    sheet.cell(row=next_row, column=6).value = date
    sheet.cell(row=next_row, column=7).value = time
    sheet.cell(row=next_row, column=4).value = semester  # Add semester
    sheet.cell(row=next_row, column=5).value = section  # Add section

    hour = int(time[:2])  # Extract the hour from the time string
    if hour in range(13, 15):
        sheet.cell(row=next_row, column=8).value = "Predictive Analysis"
    elif hour in range(11, 13):
        sheet.cell(row=next_row, column=8).value = "Image Analytics"

def attendance_system():
    global attendance_thread

    with open('student_features.pkl', 'rb') as f:
        student_features = pickle.load(f)

    registration_data = load_registration_data("")

    cap = cv2.VideoCapture(0)
    if not cap.isOpened():
        print("Error: Could not open camera.")
        return

    wb = openpyxl.load_workbook(filename="F:\\DAML\\Attendance.xlsx", data_only=True)
    sheet = wb["Sheet1"]

    current_time = datetime.datetime.now().strftime("%d-%m-%Y %H:%M:%S")
    last_attendance_times = {}

    while True:
        ret, frame = cap.read()
        if not ret:
            break

        img_small = cv2.resize(frame, (0, 0), None, 0.25, 0.25)
        imgS = cv2.cvtColor(img_small.copy(), cv2.COLOR_BGR2RGB)
        img_display = frame.copy()

        face_locations = face_recognition.face_locations(imgS)
        face_encodings = face_recognition.face_encodings(imgS, face_locations)

        for encodeFace, faceLoc in zip(face_encodings, face_locations):
            name = recognize_face(encodeFace, student_features)

            y1, x2, y2, x1 = faceLoc
            y1, x2, y2, x1 = y1 * 4, x2 * 4, y2 * 4, x1 * 4
            cv2.rectangle(img_display, (x1, y1), (x2, y2), (0, 255, 0), 2)
            cv2.putText(img_display, name, (x1, y1 - 10), cv2.FONT_HERSHEY_SIMPLEX, 0.75, (0, 255, 0), 2)

            if name in registration_data:
                reg_no = registration_data[name]['reg_no']
                section = registration_data[name]['section']
                semester = registration_data[name]['semester']
                write_to_excel(name, reg_no, current_time, sheet, last_attendance_times, semester, section)

        cv2.imshow("Webcam Feed", img_display)

        if cv2.waitKey(1) & 0xFF == 27:  # Press 'Esc' to stop
            break

    wb.save("F:\\DAML\\Attendance.xlsx")
    cap.release()
    cv2.destroyAllWindows()

@app.route('/start-attendance', methods=['POST'])
def start_attendance():
    global attendance_thread
    if attendance_thread is None or not attendance_thread.is_alive():
        attendance_thread = Thread(target=attendance_system)
        attendance_thread.start()
    return jsonify(message="Attendance system started.")

@app.route('/stop-attendance', methods=['POST'])
def stop_attendance():
    # Implement functionality to stop the attendance system if needed
    return jsonify(message="Attendance system stopped.")

@app.route('/take-attendance')
def take_attendance():
    return render_template('take_attendance.html')



# Function to extract features from the image
def extract_features(image_path, img_size=(224, 224)):
    img = cv2.imread(image_path)
    img = cv2.resize(img, img_size)
    img_rgb = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
    face_locations = face_recognition.face_locations(img_rgb)

    if face_locations:
        face_encoding = face_recognition.face_encodings(img_rgb, face_locations)[0]
        return face_encoding
    else:
        return None

# Function to save features to a file
def save_features_and_names(main_folder, output_file):
    student_features = {}

    for subfolder in os.listdir(main_folder):
        subfolder_path = os.path.join(main_folder, subfolder)
        person_features = []

        for image_file in tqdm(os.listdir(subfolder_path), desc=f"Processing {subfolder}"):
            image_path = os.path.join(subfolder_path, image_file)
            features = extract_features(image_path)

            if features is not None:
                person_features.append(features)

        # Save average features for the person
        if person_features:
            student_features[subfolder] = np.mean(person_features, axis=0)

    # Save the features to a file
    with open(output_file, 'wb') as f:
        pickle.dump(student_features, f)



@app.route('/')
def index():
    return render_template('index.html')

@app.route('/add-student', methods=['GET', 'POST'])
def add_student():
    if request.method == 'POST':
        student_data = {
            "Name": request.form['name'],
            "Regno": int(request.form['regno']),
            "Year": request.form['year'],
            "School": request.form['school'],
            "Branch": request.form['branch'],
            "Semester": request.form['semester'],
            "Section": request.form['section'],
            "Phone no": request.form['phone'],
            "DOB": request.form['dob'],
            "Blood grp": request.form['blood_group']

        }

        filename = "F:\\DAML\\Student_details.xlsx"  # Change path as needed
        print(f"Adding student to: {filename}")
        add_student_details(filename, student_data)

        # Start capturing images
        capture_and_save_images(student_data["Name"])

        return render_template('add_student.html', success=True, student_name=student_data["Name"])
      
      

    return render_template('add_student.html')

def capture_and_save_images(name):
    cam = cv2.VideoCapture(0)
    cv2.namedWindow("Capturing images")

    img_count = 1
    location = "F:\\DAML\\Project\\faces"  # Change path as needed
    folder_path = os.path.join(location, name)
    os.makedirs(folder_path, exist_ok=True)

    while True:
        ret, frame = cam.read()
        if not ret:
            print("Failed to grab frame")
            break

        cv2.imshow("Capturing images", frame)

        k = cv2.waitKey(1)
        if k % 256 == 27:  # Esc key
            print("Escape hit, closing camera")
            break
        elif k % 256 == 32:  # Space key
            for i in range(1, 28):  # Capture and augment 27 images
                img_name = f"{folder_path}/img_{i}.jpg"
                cv2.imwrite(img_name, frame)
                print(f"Screenshot {i} taken")

                img = cv2.imread(img_name)
                x = img_to_array(img)
                x = x.reshape((1,) + x.shape)

                datagen = ImageDataGenerator(
                    rotation_range=40,
                    width_shift_range=0.2,
                    height_shift_range=0.2,
                    shear_range=0.2,
                    zoom_range=0.2,
                    horizontal_flip=True,
                    fill_mode="nearest",
                )

                for batch in datagen.flow(x, batch_size=1, save_to_dir=folder_path, save_prefix=f"img_{i}_", save_format="jpg"):
                    break

    cam.release()
    cv2.destroyAllWindows()

def add_student_details(filename, student_data):
    workbook = openpyxl.load_workbook(filename, read_only=False)
    sheet = workbook.active

    max_row = sheet.max_row
    while sheet.cell(row=max_row, column=1).value is None:
        max_row -= 1
    max_row += 1

    if sheet.cell(row=max_row, column=1).value is None:
        cell_value = int(sheet.cell(row=max_row - 1, column=1).value)
        sheet.cell(row=max_row, column=1).value = cell_value + 1
    
    for col, (key, value) in enumerate(student_data.items(), start=2):
        sheet.cell(row=max_row, column=col).value = value

    workbook.save(filename)
    print("Student details appended successfully.")

@app.route('/extract-features', methods=['POST'])
def extract_features_route():
    try:
        main_folder = "F:\\DAML\\Project\\faces"  # Path to your images
        features_file = 'student_features.pkl'    # Output file for features

        # Call the function to save features
        save_features_and_names(main_folder, features_file)
        
        # Instead of flashing a message, we return a message to indicate success
        return redirect(url_for('index'))  # Redirect to index after processing
    except Exception as e:
        # In case of any error, redirect or print the error for debugging
        print(f"An error occurred: {str(e)}")
        return redirect(url_for('index'))


#check attendance
# Function to fetch and display student data from the Excel file
def fetch_and_display_student_data(excel_file,semester, section, date, start_time=None, end_time=None, subject=None):
    try:
        workbook = openpyxl.load_workbook(excel_file)
        sheet = workbook.active  # Assuming the data is in the first sheet

        present_students = []

        for row in sheet.iter_rows(min_row=2, values_only=True):  # Start from row 2, assuming header row
            if row[3] == semester:
                if row[4] == section:
                    if row[5]==date:
                        if (start_time and end_time) is None or row[6] >= start_time and row[6] < end_time:
                            if subject is None or row[7] == subject:
                                present_students.append({
                                    "name": row[1],
                                    "registration_number": row[2]
                                })

        return present_students

    except Exception as e:
        print(f"Error: {e}")
        return []

# Route to render the Check Attendance page
@app.route('/check_attendance', methods=['GET'])
def check_attendance_page():
    return render_template('check_attendance.html')



# Route to handle the form submission
@app.route('/analyze', methods=['POST'])
def analyze_attendance():
    data = request.get_json()
    date = data.get('date')
    semester = data.get('semester')
    section = data.get('section')
    start_time = data.get('startTime')
    end_time = data.get('endTime')
    subject = data.get('subject')

    wb = openpyxl.load_workbook(filename="F:\\DAML\\Attendance.xlsx", data_only=True)
    sheet = wb["Sheet1"]

    filtered_attendance = []

    # Iterate through the sheet rows and apply the filters
    for row in sheet.iter_rows(min_row=2, values_only=True):  # Skip the header row
        name, reg_no, semester_in_sheet, section_in_sheet, date_in_sheet, time_in_sheet, subject_in_sheet = row[1:8]

        # Apply filters
        if semester and semester != semester_in_sheet:
            continue
        if section and section != section_in_sheet:
            continue
        if date and date != date_in_sheet:
            continue
        if start_time and end_time:
            if not (start_time <= time_in_sheet <= end_time):
                continue
        if subject and subject != subject_in_sheet:
            continue

        # Append filtered result
        filtered_attendance.append({
            'name': name,
            'registration_number': reg_no,
            'semester': semester_in_sheet,
            'section': section_in_sheet,
            'date': date_in_sheet,
            'time': time_in_sheet,
            'subject': subject_in_sheet,
        })

    return jsonify(filtered_attendance)

        
# Route to download student details
@app.route('/download-student-details')
def download_student_details():
    path_to_file = 'F:\\DAML\\Student_details.xlsx'  
    return send_file(path_to_file, as_attachment=True)

# Route to download attendance excel sheet
@app.route('/download-attendance-excel-sheet')
def download_attendance_excel_sheet():
    path_to_file = "F:\\DAML\\Attendance.xlsx"  
    return send_file(path_to_file, as_attachment=True)

if __name__ == "__main__":
    app.run(debug=True)